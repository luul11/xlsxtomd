from openpyxl import load_workbook
import pandas as pd


# 版本：1.0
# 作者：伊玛目的门徒 (luke)  202410

def fill_merged_cells(ws):
    """处理合并单元格，将合并的单元格值填充到整个合并区域"""
    # 先收集所有合并单元格的信息
    merged_cells_info = list(ws.merged_cells.ranges)

    for merged_cell in merged_cells_info:
        # 获取合并单元格的左上角单元格的值
        top_left_cell = ws.cell(row=merged_cell.min_row, column=merged_cell.min_col)
        value = top_left_cell.value

        # 取消合并单元格
        ws.unmerge_cells(str(merged_cell))

        # 填充值到整个合并区域
        for row in ws.iter_rows(min_row=merged_cell.min_row, max_row=merged_cell.max_row,
                                min_col=merged_cell.min_col, max_col=merged_cell.max_col):
            for cell in row:
                cell.value = value


def remove_trailing_empty_rows(df):
    """去除DataFrame末尾的空行"""
    if df.isnull().values.any():  # 如果DataFrame包含空值
        last_row = df.iloc[-1]  # 获取最后一行
        while last_row.isnull().all():  # 如果最后一行全部是空值
            df = df.iloc[:-1]  # 去除最后一行
            last_row = df.iloc[-1]  # 更新最后一行
    return df


def xlsx_to_markdown(xlsx_file_path, markdown_file_path):
    print('start to read excel file')
    try:
        # 读取Excel文件
        wb = load_workbook(filename=xlsx_file_path)
        print('success to read excel file')
        ws = wb.active

        # 处理合并单元格
        fill_merged_cells(ws)

        # 将Excel数据转换为DataFrame
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        df = pd.DataFrame(data[1:], columns=data[0])  # 假设第一行是表头

        # 去除末尾的空行
        df = remove_trailing_empty_rows(df)

        # 转换为Markdown格式
        markdown_text = convert_df_to_markdown(df)

        # 保存为Markdown文件，使用UTF-8编码
        with open(markdown_file_path, 'w', encoding='utf-8') as f:
            f.write(markdown_text)
        print(f"Markdown file saved successfully: {markdown_file_path}")
    except Exception as e:
        print(f"Error reading Excel file: {e}")


def convert_df_to_markdown(df):
    """将DataFrame转换为Markdown格式"""
    markdown_text = ""
    if not df.empty:
        markdown_text += "| " + " | ".join(df.columns.astype(str)) + " |\n"
        markdown_text += "| " + " | ".join(["-" * len(str(col)) for col in df.columns.astype(str)]) + " |\n"
        for index, row in df.iterrows():
            row_data = "| " + " | ".join(map(str, row)) + " |"
            markdown_text += row_data + "\n"
    return markdown_text


def validate_file_path(file_path):
    """校验文件路径是否以.xlsx结尾"""
    if not file_path.lower().endswith('.xlsx'):
        return False
    return True


def main():
    # 提示用户输入Excel文件路径
    while True:
        xlsx_file_path = input("Please enter the path to the Excel file (.xlsx,e.g. test.xlsx): ")
        if validate_file_path(xlsx_file_path):
            break
        else:
            print("Invalid file path. Please make sure the file has a .xlsx extension.")

    markdown_file_path = xlsx_file_path.rsplit('.', 1)[0] + '.md'

    # 调用函数进行转换
    xlsx_to_markdown(xlsx_file_path, markdown_file_path)


if __name__ == "__main__":
    print ('版本：1.0  作者：伊玛目的门徒 (luke)')
    print ('project_URL:https://github.com/luul11/xlsxtomd')
    main()